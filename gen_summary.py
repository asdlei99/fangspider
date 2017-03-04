# -*- coding: utf-8 -*-

import glob
import sqlite3
import msgpack
import os
import io
import re
from argparse import ArgumentParser

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

LINK_FONT = Font(name='Microsoft YaHei UI', size=11, bold=False, italic=True, vertAlign=None, underline='single', strike=False, color='000000FF')

PICK_NUMBER = re.compile('[\\d\\.]+')

BUY_TABLE_NAME = 'buy'
DATA_PATH = 'data'
DATABASE_FILENAME = 'collection.sqlite'

PLOT_BLACKLIST = set()

if __name__ == "__main__":
    parser = ArgumentParser(add_help=True)

    config = {
        'version': '1.0.0.0'
    }

    parser.add_argument("-v", "--version", action="store_true", help="show version and exit", dest="version")
    parser.add_argument("-o", "--output", action="store", help="generated file name", dest="output", default=os.path.join(DATA_PATH, 'summary.xlsx'))
    parser.add_argument("-f", "--database-file", action="store", help="data source file", dest="data", default=os.path.join(DATA_PATH, DATABASE_FILENAME))
    parser.add_argument("-d", "--date", action="store", help="date to generate", dest="date", default=None)
    parser.add_argument("-s", "--show-date", action="store_true", help="show all available dates", dest="show_date", default=False)
    parser.add_argument("-p", "--plot-blacklist", action="store", help="where to load plot blacklist", dest="plot_blacklist", default=None)
    parser.add_argument("-n", "--destination-number", action="store", help="destination number showed in excel", dest="destination_number", type=int, default=2)

    (options, left_args) = parser.parse_known_args()

    if options.version:
        print(config['version'])
        exit(0)

    def print_help_msg(err_code):
        parser.print_help()
        exit(err_code)

    if not os.path.exists(options.data):
        print_help_msg(-1)

    if options.date is None and not options.show_date:
        print_help_msg(-1)

    COLUMNS = [
        ['数据源', 'source'],
        ['标题', 'title'],
        ['总价（万）', 'money', 'number'],
        ['单价（/㎡）', 'price', 'number'],
        ['面积（㎡）', 'area'],
        ['户型', 'room_type'],
        ['年代', 'year'],
        ['层数', 'layer'],
        ['小区', 'plot'],
        ['地址', 'address'],
        ['位置坐标', 'location'],
        ['驾车距离（米）', 'driving_distance', 'list', options.destination_number],
        ['驾车时间（分）', 'driving_duration', 'list', options.destination_number],
        ['公交距离（米）', 'integrated_distance', 'list', options.destination_number],
        ['公交时间（分）', 'integrated_duration', 'list', options.destination_number],
        ['链接', 'link', 'url'],
        ['图片', 'img', 'url'],
        ['附加信息', 'flags'],
        ['朝向', 'toward'],
        ['发布者已售数量', 'deal_count', 'number'],
        ['发布者房源数', 'publish_count', 'number']
    ]

    if options.plot_blacklist is not None:
        if os.path.exists(options.plot_blacklist):
            pbfd = open(options.plot_blacklist, 'r')
            for black_plot in pbfd.readlines():
                plot_name = black_plot.strip()
                if len(plot_name) > 0 and plot_name not in PLOT_BLACKLIST:
                    PLOT_BLACKLIST.add(plot_name)
            pbfd.close()
        else:
            print('{0} is not exists'.format(options.plot_blacklist))
            print_help_msg(-1)

    sheet_name = options.date
    # 先检查有没有数据
    dbconn = sqlite3.connect(options.data)
    dbconn.text_factory = str
    dbcursor = dbconn.cursor()

    if options.show_date:
        rows = dbcursor.execute("select distinct create_date from {0}".format(BUY_TABLE_NAME))
        rows_data = rows.fetchall()
        print('There are {0} dates in table {1} of file {2}'.format(len(rows_data), BUY_TABLE_NAME, options.data))
        for data_element in rows_data:
            print(data_element[0])
        exit(0)

    rows = dbcursor.execute("select url,data,create_date from {0} where create_date=:create_date".format(BUY_TABLE_NAME), {'create_date': sheet_name})
    row_data = rows.fetchone()
    if row_data is None:
        print('[WARN] there is no data with date = {0} in table {1}'.format(sheet_name, BUY_TABLE_NAME))
        exit(0)

    xlsx_file_path = options.output
    if os.path.exists(xlsx_file_path):
        wb = load_workbook(xlsx_file_path)
    else:
        wb = Workbook()

    if sheet_name in wb.get_sheet_names():
        print('[INFO] remove sheet {0}'.format(sheet_name))
        wb.remove_sheet(wb.get_sheet_by_name(sheet_name))
    ws = wb.create_sheet(title=sheet_name)

    data_set = []
    for kv in COLUMNS:
        if len(kv) > 2 and kv[2] == 'list':
            for idx in range(1, kv[3] + 1):
                data_set.append('{0}-{1}'.format(kv[0], idx))
        else:
            data_set.append(kv[0])
    ws.append(data_set)

    def pick_data(data_dict, key):
        if key in data_dict:
            ret = data_dict[key]
            if isinstance(ret, str):
                return ret
            elif isinstance(ret, float) or isinstance(ret, int):
                return str(ret)
            elif isinstance(ret, list):
                return ret
            return ret.encode('utf-8')
        return ''

    def pick_number(data_dict, key):
        str_content = pick_data(data_dict, key)
        res = PICK_NUMBER.search(str_content)
        if res is not None:
            return float(res.group(0))
        return 0

    column_index = 1
    while row_data is not None:
        data_element = row_data
        try:
            data_obj = msgpack.unpackb(data_element[1], encoding='utf-8')
            # filter by some segments
            plot_name = pick_data(data_obj, 'plot')
            if len(plot_name) <= 0 or plot_name not in PLOT_BLACKLIST:
                column_index = column_index + 1
                data_set = []
                url_format = []
                column_number = 0
                for i in range(0, len(COLUMNS)):
                    kv = COLUMNS[i]
                    column_number = column_number + 1
                    if len(kv) == 2:
                        data_set.append(pick_data(data_obj, kv[1]))
                    elif kv[2] == 'number':
                        data_set.append(pick_number(data_obj, kv[1]))
                    elif kv[2] == 'url':
                        val = pick_data(data_obj, kv[1])
                        data_set.append(val)
                        if len(val) > 0:
                            url_format.append(column_number)
                    elif kv[2] == 'list':
                        val = pick_data(data_obj, kv[1])
                        for j in range(0, kv[3]):
                            if j < len(val):
                                data_set.append(val[j])
                            if j > 0:
                                column_number = column_number + 1
                ws.append(data_set)
                for url_col in url_format:
                    link_cell = ws.cell(row=column_index, column=url_col)
                    link_cell.hyperlink = link_cell.value
                    link_cell.font = LINK_FONT
                print('[INFO] generate data for {0} success'.format(data_element[0]))
        except Exception as err:
            print('[INFO] generate data for {0} failed\n{1}'.format(data_element[0], err))

        row_data = rows.fetchone()

    wb.save(xlsx_file_path)
